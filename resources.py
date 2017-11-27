import random
import openpyxl


class Partaker(object):
    def __init__(self, name, nick=''):
        self.name = name
        self.fullname = self.name
        if nick:
            self.nick = '\"' + nick + '\"'
            temp_str_list = self.name.split(' ')
            temp_str_list.insert(1, self.nick)
            self.fullname = ' '.join(temp_str_list)


class Player(Partaker):
    def __init__(self, name, nick):
        super().__init__(name, nick)


class Master(Partaker):
    def __init__(self, name, nick):
        super().__init__(name, nick)
        self.room = ''
        self.session_name = ''
        self.max_players = 2
        self.players = []


class Contest(object):

    name_names_list = ['Imię i nazwisko', 'Imię']
    nick_names_list = ['Pseudonim', 'Ksywka', 'Pseudonim/Ksywka']

    def __init__(self, name, plik_mg_nazwa='mg.xlsx', plik_gracze_nazwa='gracze.xlsx'):
        self.name = name
        self.masters_list = []
        self.players_list = []
        self.sesje = []
        self.sale = ['RPG1', 'RPG2', 'RPG3']

        self.players_file = openpyxl.load_workbook(plik_gracze_nazwa)
        self.masters_file = openpyxl.load_workbook(plik_mg_nazwa)
        self.players_sheet = self.players_file.active
        self.masters_sheet = self.masters_file.active

    def load_masters(self):
        column_titles = {}
        for cell in self.masters_sheet[1]:
            if cell.value in self.name_names_list:
                column_titles['names'] = cell.column
            if cell.value in self.nick_names_list:
                column_titles['nicks'] = cell.column
        # print(dict)
        for row in self.masters_sheet.iter_rows(min_row=2):
            temp_name = ''
            temp_nick = ''
            for cell in row:
                if cell.column is column_titles['names']:
                    temp_name = cell.value
                if cell.column is column_titles['nicks']:
                    temp_nick = cell.value
            self.masters_list.append(Master(temp_name, temp_nick))
        # for mg in self.masters_list:
        #     print(mg.fullname)

    def load_players(self):
        column_titles = {}
        for cell in self.players_sheet[1]:
            if cell.value in self.name_names_list:
                column_titles['names'] = cell.column
            if cell.value in self.nick_names_list:
                column_titles['nicks'] = cell.column
        # print(dict)
        for row in self.players_sheet.iter_rows(min_row=2):
            temp_name = ''
            temp_nick = ''
            for cell in row:
                if cell.column is column_titles['names']:
                    temp_name = cell.value
                if cell.column is column_titles['nicks']:
                    temp_nick = cell.value
            self.players_list.append(Player(temp_name, temp_nick))
        # for item in self.players_list:
        #     print(item.fullname)

    def populate_sessions(self):
        session_id = 0
        for master in self.masters_list:
            master.room = self.sale[session_id]
            session_id = session_id + 1
        temp_players = self.players_list
        random.shuffle(temp_players)
        full_sessions = 0
        while temp_players and full_sessions < len(self.masters_list):
            for master in self.masters_list:
                if len(master.players) < master.max_players:
                    try:
                        master.players.append(temp_players.pop())
                    except IndexError:
                        print('Brak graczy!')
                else:
                    full_sessions = full_sessions + 1
